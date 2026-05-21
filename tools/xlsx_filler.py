"""Format-preserving xlsx template filler.

Why this exists
---------------
``openpyxl`` is the obvious choice for Excel work in Python, but it
**re-serialises** every part of the workbook on save. That re-serialisation
silently drops:

  * floating drawings / shapes (the angled lines that subsidy forms use
    for "斜線が入る記載例" markings)
  * conditional formatting tied to specific shared-string indices
  * defined names with complex scopes
  * data validation lists
  * embedded images positioned by anchor
  * cell-level locked/protected attributes

For Japanese subsidy applications, the templates are often hand-tuned with
exactly those features, and losing them means the submitted file is
visibly wrong. The official submission portal sometimes rejects them.

What this module does
---------------------
Treats the .xlsx as the ZIP archive it actually is and edits only the
text content inside ``sharedStrings.xml`` and the per-sheet inline string
nodes. Every other XML part is copied byte-for-byte from the original.
The result is a workbook that's indistinguishable from the template
except for the substituted text — drawings, formatting, formulas, and
defined names all survive.

Usage::

    fill_xlsx_template(
        template_path="templates/jizoku_19/経費明細書.xlsx",
        out_path="demo/output/経費明細書.xlsx",
        substitutions={
            "applicant_name": "株式会社サンプル珈琲",
            "expense_1_category": "ウェブサイト関連費",
            "expense_1_amount": "500000",
            ...
        },
    )

Placeholder syntax is the same as the .docx filler: ``{{key}}`` with
optional whitespace.
"""
from __future__ import annotations

import logging
import re
import shutil
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

logger = logging.getLogger(__name__)


_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z0-9_\-\.]+)\s*\}\}")

# XML files inside an .xlsx that hold user-visible text
_TEXT_BEARING_FILES_PREFIXES = (
    "xl/sharedStrings.xml",
    "xl/worksheets/sheet",
)


def fill_xlsx_template(
    template_path: str | Path,
    out_path: str | Path,
    substitutions: dict[str, str],
) -> dict:
    """Open template, substitute placeholders in text parts, save preserving all else.

    Returns a report::

        {
          "replaced": <total replacement count across the workbook>,
          "files_touched": [<list of XML parts that had at least one replacement>],
          "missing_keys": [<placeholders left unresolved>],
          "unique_keys_used": <distinct substitution keys actually used>,
        }
    """
    template_path = Path(template_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists():
        out_path.unlink()

    # Track per-key usage and per-file replacements
    keys_used: dict[str, int] = {}
    files_touched: list[str] = []

    with zipfile.ZipFile(template_path, "r") as zin, zipfile.ZipFile(
        out_path, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if _is_text_bearing(item.filename):
                new_data, replaced_here = _substitute_in_xml_bytes(
                    data, substitutions, keys_used
                )
                if replaced_here > 0:
                    files_touched.append(item.filename)
                data = new_data
            # Use ZipInfo to preserve compression and date_time
            zout.writestr(item, data)

    # Scan for any unresolved placeholders (read-back)
    missing = _scan_unresolved_placeholders(out_path)

    return {
        "replaced": sum(keys_used.values()),
        "unique_keys_used": len(keys_used),
        "files_touched": files_touched,
        "missing_keys": sorted(missing),
    }


def _is_text_bearing(name: str) -> bool:
    return any(name.startswith(prefix) for prefix in _TEXT_BEARING_FILES_PREFIXES)


def _substitute_in_xml_bytes(
    data: bytes,
    substitutions: dict[str, str],
    keys_used: dict[str, int],
) -> tuple[bytes, int]:
    """Replace {{key}} placeholders that appear inside XML text content.

    We decode to text, run regex substitution, and re-encode. Because the
    XML is well-formed and placeholders only ever appear inside text nodes
    (between ``<t>...</t>`` in sharedStrings.xml or in inline ``<is><t>``
    blocks in sheet xml), substituting at the text level cannot break the
    XML structure as long as the replacement is properly XML-escaped.
    """
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        return data, 0

    if "{{" not in text:
        return data, 0

    local_count = 0

    def _replace(match: re.Match) -> str:
        nonlocal local_count
        key = match.group(1)
        if key not in substitutions:
            return match.group(0)  # leave placeholder so we can scan and warn
        value = substitutions[key]
        keys_used[key] = keys_used.get(key, 0) + 1
        local_count += 1
        return xml_escape(str(value))

    new_text = _PLACEHOLDER_RE.sub(_replace, text)
    return new_text.encode("utf-8"), local_count


def _scan_unresolved_placeholders(xlsx_path: Path) -> set[str]:
    leftovers: set[str] = set()
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        for item in zin.infolist():
            if not _is_text_bearing(item.filename):
                continue
            data = zin.read(item.filename)
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                continue
            leftovers.update(_PLACEHOLDER_RE.findall(text))
    return leftovers


# ---------------------------------------------------------------------------
# Helpers for building a sample template programmatically (demo / tests)
# ---------------------------------------------------------------------------


def write_sample_xlsx_template(out_path: str | Path) -> Path:
    """Generate a minimal .xlsx template with placeholders.

    Uses openpyxl ONLY for the initial template authoring (not for
    fill-time round-trip). The filler itself never touches openpyxl.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "経費明細書"

    # Header
    ws["A1"] = "経費明細書（テンプレート）"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:E1")

    ws["A3"] = "申請事業者名"
    ws["B3"] = "{{ applicant_name }}"
    ws["A4"] = "代表者"
    ws["B4"] = "{{ representative }}"

    headers = ["No.", "経費区分", "経費項目", "金額（円）", "備考"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
        side = Side(border_style="thin", color="000000")
        cell.border = Border(top=side, bottom=side, left=side, right=side)

    for i in range(1, 6):
        row = 6 + i
        ws.cell(row=row, column=1, value=str(i))
        ws.cell(row=row, column=2, value=f"{{{{ expense_{i}_category }}}}")
        ws.cell(row=row, column=3, value=f"{{{{ expense_{i}_item }}}}")
        ws.cell(row=row, column=4, value=f"{{{{ expense_{i}_amount }}}}")
        ws.cell(row=row, column=5, value=f"{{{{ expense_{i}_note }}}}")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = Border(
                top=Side(border_style="thin", color="999999"),
                bottom=Side(border_style="thin", color="999999"),
                left=Side(border_style="thin", color="999999"),
                right=Side(border_style="thin", color="999999"),
            )

    ws["C12"] = "合計"
    ws["C12"].font = Font(bold=True)
    ws["C12"].alignment = Alignment(horizontal="right")
    ws["D12"] = "{{ expense_total }}"
    ws["D12"].font = Font(bold=True)

    ws["A14"] = "補助金申請額"
    ws["B14"] = "{{ subsidy_amount }}"
    ws["A15"] = "自己負担額"
    ws["B15"] = "{{ self_funding }}"

    # Column widths
    widths = {"A": 6, "B": 22, "C": 32, "D": 14, "E": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _copy_xlsx(src: Path, dst: Path) -> None:
    """Byte-identical copy preserving binary content."""
    shutil.copy2(src, dst)
