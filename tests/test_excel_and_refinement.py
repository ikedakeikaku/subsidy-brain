"""Tests for the Excel filler, adoption estimator, and refinement loop."""
from __future__ import annotations

import asyncio
import sys
import zipfile
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


# ---------------------------------------------------------------------------
# Excel filler — format preservation is critical
# ---------------------------------------------------------------------------


def test_xlsx_filler_replaces_placeholders(tmp_path: Path) -> None:
    from tools.xlsx_filler import fill_xlsx_template, write_sample_xlsx_template

    template = write_sample_xlsx_template(tmp_path / "template.xlsx")
    out = tmp_path / "filled.xlsx"

    report = fill_xlsx_template(
        template,
        out,
        {
            "applicant_name": "テスト株式会社",
            "representative": "山田太郎",
            "expense_1_category": "ウェブサイト関連費",
            "expense_1_item": "EC構築",
            "expense_1_amount": "500,000",
            "expense_1_note": "",
            "expense_2_category": "",
            "expense_2_item": "",
            "expense_2_amount": "",
            "expense_2_note": "",
            "expense_3_category": "",
            "expense_3_item": "",
            "expense_3_amount": "",
            "expense_3_note": "",
            "expense_4_category": "",
            "expense_4_item": "",
            "expense_4_amount": "",
            "expense_4_note": "",
            "expense_5_category": "",
            "expense_5_item": "",
            "expense_5_amount": "",
            "expense_5_note": "",
            "expense_total": "500,000",
            "subsidy_amount": "333,000",
            "self_funding": "167,000",
        },
    )
    assert report["missing_keys"] == []
    assert report["replaced"] >= 7

    wb = load_workbook(out)
    ws = wb.active
    assert ws["B3"].value == "テスト株式会社"
    assert ws["D7"].value == "500,000"


def test_xlsx_filler_does_not_touch_non_text_parts(tmp_path: Path) -> None:
    """Strict format preservation: only sharedStrings / sheets are modified.

    Anything under xl/styles.xml, xl/drawings/, xl/theme/, xl/_rels/ etc.
    must be byte-identical to the template.
    """
    from tools.xlsx_filler import fill_xlsx_template, write_sample_xlsx_template

    template = write_sample_xlsx_template(tmp_path / "template.xlsx")
    out = tmp_path / "filled.xlsx"

    fill_xlsx_template(
        template,
        out,
        {f"expense_{i}_{f}": "" for i in range(1, 6) for f in
         ("category", "item", "amount", "note")}
        | {
            "applicant_name": "x",
            "representative": "y",
            "expense_total": "0",
            "subsidy_amount": "0",
            "self_funding": "0",
        },
    )

    sensitive_prefixes = (
        "xl/styles.xml",
        "xl/theme/",
        "xl/_rels/",
        "_rels/",
        "docProps/",
    )

    with zipfile.ZipFile(template, "r") as ztpl, zipfile.ZipFile(out, "r") as zout:
        tpl_names = set(ztpl.namelist())
        out_names = set(zout.namelist())
        assert tpl_names == out_names, "file list must be identical"
        for name in tpl_names:
            if not any(name.startswith(p) for p in sensitive_prefixes):
                continue
            assert ztpl.read(name) == zout.read(name), (
                f"{name} should be byte-identical between template and output"
            )


# ---------------------------------------------------------------------------
# Adoption estimator — distinct from quality_scoring
# ---------------------------------------------------------------------------


def test_adoption_estimator_returns_six_signals() -> None:
    from demo.mock_story import MOCK_STORY
    from schemas.subsidy_profile import load_profile
    from tools.adoption_estimator import estimate_adoption_probability

    company = yaml.safe_load(
        (ROOT / "demo" / "sample_company.yaml").read_text(encoding="utf-8")
    )
    profile = load_profile(ROOT / "demo" / "sample_profile.yaml")
    report = estimate_adoption_probability(profile, company, MOCK_STORY)

    expected = {
        "自社固有データ",
        "課題→施策対応",
        "加点項目活用",
        "具体数値密度",
        "文字数達成",
        "図表配置",
    }
    assert {s["name"] for s in report["signals"]} == expected
    assert 0 <= report["total"] <= 100
    # Mock story is comprehensive enough to pass.
    assert report["passed"] is True


# ---------------------------------------------------------------------------
# Refinement loop — must improve a deliberately weak draft
# ---------------------------------------------------------------------------


def test_refinement_loop_improves_weak_story() -> None:
    from schemas.subsidy_profile import load_profile
    from tools.adoption_estimator import estimate_adoption_probability
    from tools.refinement_loop import refine_until_threshold

    company = yaml.safe_load(
        (ROOT / "demo" / "sample_company.yaml").read_text(encoding="utf-8")
    )
    profile = load_profile(ROOT / "demo" / "sample_profile.yaml")

    weak = {sid: "短い" for sid in [
        "section_1_1", "section_1_2", "section_1_3",
        "section_2_1", "section_3", "section_4_2",
        "section_effect", "bonus_env_change",
    ]}
    before = estimate_adoption_probability(profile, company, weak)
    result = asyncio.run(
        refine_until_threshold(
            profile, company, weak,
            target_score=70, max_iterations=8, mode="mock",
        )
    )

    assert result["final_score"] > before["total"], result["iterations"]
    # Each iteration should have touched a different weak section.
    refined_sections = [
        it["refined_section"] for it in result["iterations"]
        if it["refined_section"] is not None
    ]
    assert len(refined_sections) == len(set(refined_sections)), (
        "loop should not refine the same section twice"
    )


# ---------------------------------------------------------------------------
# Preset registry loads
# ---------------------------------------------------------------------------


def test_jizoku_19_preset_loads() -> None:
    from schemas.subsidy_profile import load_profile
    from schemas.subsidy_registry import YamlSubsidyRegistry

    reg = YamlSubsidyRegistry(ROOT / "presets" / "jizoku_19.yaml")
    prog = reg.get("jizoku_19")
    assert prog is not None
    assert prog.canonical_name.startswith("小規模事業者持続化補助金")
    assert prog.round_number == 19
    assert len(prog.forms) >= 4
    assert any(f.local_path.endswith(".xlsx") for f in prog.forms), (
        "preset should reference at least one xlsx form"
    )

    profile = load_profile(ROOT / "presets" / "jizoku_19_profile.yaml")
    assert profile.total_target_chars >= 7_000  # 持続化補助金 has ~10pp of writing
